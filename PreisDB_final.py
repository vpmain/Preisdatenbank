import sys
import statistics
import logging
from email.policy import strict

from dateutil import rrule

import pyodbc
import getpass

from babel.numbers import format_number, format_decimal, format_percent
from datetime import datetime, date
from qtpy import QtWidgets, QtGui, QtCore
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QMenu, QComboBox, QDateEdit, QDateTimeEdit, QLabel, QWidget, QShortcut
from PyQt5 import QtSql, QtWidgets
from PyQt5.QtSql import QSqlDatabase, QSqlQueryModel, QSqlQuery

from PreisDB.preisdb import Ui_PreisDB
from PreisDB.splashscreen import Ui_SplashScreen
from PreisDB.indexrechner import Ui_IndexRechner

from openpyxl import Workbook, load_workbook

# ==> GLOBALS
counter = 0
SERVER_NAME = 'SQLAO-AG06L\APPS'
DATABASE_NAME = 'PreisDB'
SERVER_NAME2 = 'SQLSPSRV01\SP2K16DEV'
DATABASE_NAME2 = 'BE_TEST'
programmname = "V+P-Apps/Preisdatenbank"
programmversion = "1.2"
path_templates_LV = r"\\firma.local\dfs\Firmenstandard\Preisdatenbank\templates\Template_LV_Export.xlsx"
path_templates_Kennwerte = r"\\firma.local\dfs\Firmenstandard\Preisdatenbank\templates\Template_Kennwerte_Export.xlsx"
path_templates_EHP_Calc = r"\\firma.local\dfs\Firmenstandard\Preisdatenbank\templates\Template_EHP_Calc_Export.xlsx"
company_icon = r"\\firma.local\dfs\Firmenstandard\Preisdatenbank\icon\favicon.ico"
lg = [
    "-",
    "01 - Baustellengemeinkosten",
    "02 - Abbruch",
    "03 - Roden, Baugrube, Sicherungen u. Tiefgründungen",
    "06 - Aufschließung, Infrastruktur",
    "07 - Beton- und Stahlbetonarbeiten",
    "08 - Mauerarbeiten",
    "09 - Versetzarbeiten",
    "10 - Putz",
    "11 - Estricharbeiten",
    "12 - Abdichtungen bei Betonflächen und Wänden",
    "13 - Außenanlagen",
    "14 - Besondere Instandsetzungsarbeiten",
    "15 - Schlitze, Durchbrüche, Sägen u. Bohren",
    "16 - Fertigteile",
    "18 - Winterbauarbeiten",
    "19 - Baureinigung",
    "20 - Regieleistungen",
    "21 - Dachabdichtungsarbeiten",
    "22 - Dachdeckerarbeiten",
    "23 - Bauspenglerarbeiten",
    "24 - Fliesen- und Plattenlegearbeiten",
    "25 - Sicherheits- und Schutzmaßnahmen",
    "26 - (Asphaltarbeiten)",
    "27 - Terrazzoarbeiten",
    "28 - Natursteinarbeiten",
    "29 - Kunststeinarbeiten",
    "30 - Schließanlagen",
    "31 - Metallbauarbeiten",
    "32 - Konstruktiver Stahlbau",
    "34 - Verglaste Rohrrahmenelemente",
    "35 - System-Abgasanlagen",
    "36 - Zimmermeisterarbeiten",
    "37 - Tischlerarbeiten",
    "38 - Holzfußböden",
    "39 - Trockenbauarbeiten",
    "42 - Glasarbeiten",
    "43 - Türsysteme (Elemente)",
    "44 - Wärmedämmverbundsystem (WDVS)",
    "45 - Beschichtungen auf Holz und Metall",
    "46 - Beschichtungen auf Mauerwerk, Putz und Beton",
    "47 - Tapetenarbeiten",
    "49 - Beschichtungen von Betonböden",
    "50 - Klebearbeiten für Boden und Wandbeläge",
    "51 - Fenster und Fenstertüren aus Holz",
    "52 - Fenster und Fenstertüren aus Aluminium",
    "53 - Fenster und Fenstertüren aus Kunststoff",
    "54 - Fenster und Fenstertüren aus Holz-Alu",
    "55 - Sanierung von Fenster und Türen aus Holz",
    "56 - Dachflächenfenster, Lichtkuppeln, Lichtbänder",
    "57 - Bewegliche Anschlüsse von Fenstern",
    "58 - Gartengestaltung und Landschaftsbau",
    "59 - Sportanlagen im Freien",
    "61 - Sporthallenausbau",
    "65 - Toranlagen in Gebäuden",
    "66 - (Beschriftung und Beschilderung))",
    "67 - Pfosten-Riegel-Fassaden aus Alu",
    "68 - Vorgehängte Hinterlüftete Fassaden",
    "69 - (Leitsystem)",
    "70 - Haustechnik - Allgemein",
    "71 - Haustechnik - Förderanlagen",
    "72 - Haustechnik - Wärmeversorgungsanlagen",
    "73 - Haustechnik - Klima-/Lüftungsanlagen",
    "74 - Haustechnik - Sanitär und Gasanlagen",
    "75 - Haustechnik - Starkstromanlagen",
    "76 - Haustechnik - Schwachstromanlagen",
    "77 - Haustechnik - Gebäudeautomation",
    "78 - Haustechnik - Betriebseinrichtungen",
    "79 - Haustechnik - Außenanlagen",
    "80 - (Einrichtung)",
    "81 - (Änderung Gesamtprojekt)",
    "82 - (Provisorien)",
    "90 - Schutzraumeinbauteile und Einrichtungen",
    "93 - (GU-Zuschlag)",
    "94 - (Valorisierung)",
    "96 - (Aufschließung)",
    "97 - (Planungsleistungen)",
    "98 - (Projektnebenleistungen)",
]

app = QtWidgets.QApplication(sys.argv)


## Connection über pyodbc
def db_connection():
    server = SERVER_NAME
    database = DATABASE_NAME
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=' + server + ';'
                                               'Database=' + database + ';'
                                                                        'Trusted_Connection=yes;')
    return conn


## Connection über QtSql
def createConnection():
    ## CoonectionString definieren / Variablen in den Globals
    connString = f'DRIVER={{SQL Server}};' \
                 f'SERVER={SERVER_NAME};' \
                 f'DATABASE={DATABASE_NAME};' \
                 f'Trusted_Connection=yes;'

    global db
    global msg
    db = QtSql.QSqlDatabase.addDatabase('QODBC')
    db.setDatabaseName(connString)

    if db.open():
        print('connect to SQL Server successfully')
        msg = "Datenbankverbindung erfolgreich"
        return True
    else:
        print('connection failed')
        msg = "Datenbankverbindung fehlgeschlagen"
        return False


def createConnection2():
    ## CoonectionString definieren / Variablen in den Globals
    connString = f'DRIVER={{SQL Server}};' \
                 f'SERVER={SERVER_NAME2};' \
                 f'DATABASE={DATABASE_NAME2};' \
                 f'Trusted_Connection=yes;'

    global db2
    global msg2
    db2 = QtSql.QSqlDatabase.addDatabase('QODBC')
    db2.setDatabaseName(connString)

    if db2.open():
        print('connect to SQL Server successfully')
        msg2 = "Datenbankverbindung erfolgreich"
        return True
    else:
        print('connection failed')
        msg2 = "Datenbankverbindung fehlgeschlagen"
        return False


def string_to_float(value: str, rounded: int):
    new_value = round(float(value.replace(".", "").replace(",", ".")), rounded)
    return new_value


def float_to_string(value: str):
    new_value = format_decimal(value, format='#,##0.00', locale='de_DE')
    return new_value


class MySqlModel(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):  # =Qt.DisplayRole
        if role == Qt.TextAlignmentRole:
            # Perform per-type checks and render accordingly.
            if index.column() >= 0 and index.column() <= 1:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 2:
                return Qt.AlignVCenter + Qt.AlignRight

            if index.column() == 3:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() >= 4 and index.column() <= 7:
                return Qt.AlignVCenter + Qt.AlignRight

            if index.column() >= 8 and index.column() <= 9:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() >= 10 and index.column() <= 11:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() >= 12 and index.column() <= 14:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() >= 15 and index.column() <= 18:
                return Qt.AlignVCenter + Qt.AlignRight

            # Default (anything not captured above: e.g. int)
            return Qt.AlignVCenter + Qt.AlignLeft

        if role == Qt.FontRole:
            if index.column() == 5:
                font = QtGui.QFont()
                font.setBold(True)
                return font

        return QSqlQueryModel.data(self, index, role)


class MySqlModel2(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):  # =Qt.DisplayRole
        if role == Qt.TextAlignmentRole:
            # Perform per-type checks and render accordingly.
            if index.column() == 0:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 1:
                return Qt.AlignVCenter + Qt.AlignRight

            # Default (anything not captured above: e.g. int)
            return Qt.AlignVCenter + Qt.AlignLeft

        return QSqlQueryModel.data(self, index, role)


class SQLModelIndexVerwaltung(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):  # =Qt.DisplayRole
        if role == Qt.TextAlignmentRole:
            # Perform per-type checks and render accordingly.
            if index.column() == 0:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 1:
                return Qt.AlignVCenter + Qt.AlignRight

            # Default (anything not captured above: e.g. int)
            return Qt.AlignVCenter + Qt.AlignLeft

        return QSqlQueryModel.data(self, index, role)


class SqlModelKennwerte(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):  # =Qt.DisplayRole
        if role == Qt.TextAlignmentRole:
            # Perform per-type checks and render accordingly.
            if index.column() == 0:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 1:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 2:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 3:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 4:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() >= 5 and index.column() <= 6:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() >= 7:
                return Qt.AlignVCenter + Qt.AlignRight

            # Default (anything not captured above: e.g. int)
            return Qt.AlignVCenter + Qt.AlignLeft

        return QSqlQueryModel.data(self, index, role)


class SqlModelKennwerteDetails(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):  # =Qt.DisplayRole
        if role == Qt.TextAlignmentRole:
            # Perform per-type checks and render accordingly.
            if index.column() == 0:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 1:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 2:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 3:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 4:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 5:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() >= 6:
                return Qt.AlignVCenter + Qt.AlignRight

            # Default (anything not captured above: e.g. int)
            return Qt.AlignVCenter + Qt.AlignLeft

        return QSqlQueryModel.data(self, index, role)


class SqlModelKennwerteDetailsLVPos(QSqlQueryModel):
    def data(self, index, role=Qt.DisplayRole):  # =Qt.DisplayRole
        if role == Qt.TextAlignmentRole:
            # Perform per-type checks and render accordingly.
            if index.column() == 0:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 1:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 2:
                return Qt.AlignVCenter + Qt.AlignLeft

            if index.column() == 3:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() == 4:
                return Qt.AlignVCenter + Qt.AlignRight

            if index.column() == 5:
                return Qt.AlignVCenter + Qt.AlignHCenter

            if index.column() >= 6 and index.column() <= 13:
                return Qt.AlignVCenter + Qt.AlignRight

            if index.column() >= 14:
                return Qt.AlignVCenter + Qt.AlignHCenter

            # Default (anything not captured above: e.g. int)
            return Qt.AlignVCenter + Qt.AlignLeft

        return QSqlQueryModel.data(self, index, role)


class CheckScreen(QWidget):
    def __init__(self):
        super().__init__()
        self.setFixedSize(128, 128)
        self.setWindowFlags(Qt.WindowStaysOnTopHint | QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        self.label_animation = QLabel(self)
        self.label_animation.setPixmap(QPixmap(r"\\firma.local\dfs\Firmenstandard\Preisdatenbank\icon\check.png"))

        timer = QTimer(self)
        self.showWindow()
        timer.singleShot(1000, self.hideWindow)

    def showWindow(self):
        self.show()

    def hideWindow(self):
        self.close()


class SplashScreen(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.ui = Ui_SplashScreen()
        self.ui.setupUi(self)

        self.setWindowIcon(QtGui.QIcon(company_icon))

        # self.ui.splashTitlePic.setPixmap(QPixmap(".img\\Klammer.png"))
        self.ui.splashTitlePic.setPixmap(QPixmap(r"\\firma.local\dfs\Firmenstandard\Preisdatenbank\icon\Klammer.png"))

        ## REMOVE TITLE BAR
        self.setWindowFlags(Qt.WindowStaysOnTopHint | QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        ## DB VALUES

        createConnection()

        ## Projektanzahl
        self.qry1 = """
            SELECT Count(DISTINCT GZ)
            FROM ProjekteView
        """
        self.query = QtSql.QSqlQuery()
        self.query.exec(self.qry1)
        self.query.first()
        self.string = '{:,}'.format(self.query.value(0)).replace(",", ".")

        ############################################################################

        ## Positonsanzahl
        self.qry2 = """
            SELECT Count(*)
            FROM GewerkeViewOhneBieter
        """
        self.query.exec(self.qry2)
        self.query.first()
        self.string2 = '{:,}'.format(self.query.value(0)).replace(",", ".")

        ############################################################################

        ## Bieterpreisanzahl
        self.qry3 = """
            SELECT Count(GZ)
            FROM GewerkeView
        """
        self.query.exec(self.qry3)
        self.query.first()
        self.string3 = '{:,}'.format(self.query.value(0)).replace(",", ".")

        ############################################################################

        ## QTIMER ==> Start
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.progress)

        ## TIMER IN MILLISECONDS
        self.timer.start(35)

    def progress(self):
        global counter

        # SET VALUE TO PROGRESS BAR
        self.ui.splashProgressBar.setValue(counter)

        # COUNTER = 1
        if counter == 1:
            self.ui.splashText.setText(msg)

        # COUNTER = 25
        if counter == 25:
            self.ui.splashText.setText(self.string + " Projekte werden geladen...")

        # COUNTER = 50
        if counter == 50:
            self.ui.splashText.setText(self.string2 + " Positionen werden geladen...")

        # COUNTER = 75
        if counter == 75:
            self.ui.splashText.setText(self.string3 + " Bieterpreise werden geladen...")

        # CLOSE SPLASHSCREEN AND OPEN APP
        if counter > 100:
            # STOP TIMER
            self.timer.stop()

            # SHOW MAIN WINDOW
            self.main = MainWindow()
            self.main.show()

            # CLOSE SPLASHSCREEN
            self.close()

        # INCREASE COUNTER
        counter += 1


class MainWindow(QtWidgets.QMainWindow):

    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)

        self.ui = Ui_PreisDB()
        self.ui.setupUi(self)

        ## Partner-Dialogs
        self.partnerDialogCalc = IndexRechner(self)
        # self.partnerDialogEHPCalc = EHPCalc(self)
        # self.partnerDialogEHPCalcDetail = EHPCalcDetail(self)
        # self.partnerDialogLVDetail = LVPosDetail(self)

        ## Window options
        self.setWindowTitle("Preisdatenbank [Version: " + programmversion + "]")
        # self.setWindowIcon(QtGui.QIcon(".icon\\favicon.ico"))
        self.setWindowIcon(QtGui.QIcon(company_icon))
        self.setWindowIconText("logo")
        self.showMaximized()
        self.ui.lblTitle.setText("Suche: LV-Positionen")
        self.ui.lblTitleKennBau.setText("Kennwerte Bau")
        self.ui.lblTitleVerwaltung.setText("Verwaltung")

        ## Start functions
        # self.table_settings() #Set Columntitles
        self.set_controls_lv()  # Set controlelements
        self.set_controls_kenn()
        self.set_controls_index()
        self.ui.grdView.customContextMenuRequested.connect(self.generate_menu_lv)  # Generate right-click menu in table
        self.ui.grdViewKenn.customContextMenuRequested.connect(self.generate_menu_kenn)
        self.ui.grdViewDetailKenn.customContextMenuRequested.connect(self.generate_menu_kenn_detail)
        self.show_lv()
        self.show_lv_main()

        ## Definie Actionbar
        # self.ui.actIndexRechner.setIcon(QtGui.QIcon(".icon\\calculator.png"))
        self.ui.actIndexRechner.setIcon(QtGui.QIcon(r"\\firma.local\dfs\Firmenstandard\Preisdatenbank\icon\calculator.png"))
        self.ui.actIndexRechner.triggered.connect(self.open_index)
        self.ui.actIndexRechner.setShortcut(QKeySequence("Ctrl+I"))

        self.ui.actReset.triggered.connect(self.reset_controls_lv)
        self.ui.actReset.triggered.connect(self.reset_controls_kenn)
        self.ui.actReset.setShortcut(QKeySequence("Ctrl+R"))

        self.ui.actLV_Position_Suche.triggered.connect(self.show_lv)
        self.ui.actKennwerteBau.triggered.connect(self.show_kennwerte_bau_main)
        self.ui.actVerwaltung.triggered.connect(self.show_verwaltung)
        #
        # self.ui.actEHPCalc.triggered.connect(self.open_ehp_calc)

        # self.ui.actKennwerteBau.setVisible(False)
        self.ui.actionProjekt_bersicht.setVisible(False)

        self.ui.actExportXLSXselected.triggered.connect(self.export_excel_selected)

        ## Shortcuts
        self.execute_query = QShortcut(QKeySequence('Return'), self)
        self.execute_query.activated.connect(self.db_fill_table)
        self.execute_query = QShortcut(QKeySequence('Enter'), self)
        self.execute_query.activated.connect(self.db_fill_table)

        ## Definie Button Actions LV-Suche
        self.ui.cmdStartQuery.clicked.connect(self.start_query)
        self.ui.cmdCalc.clicked.connect(self.mw_calc_lv)
        self.ui.cmdDetailBack.clicked.connect(self.show_lv_main)
        self.ui.cmdUnselect.setVisible(False)

        ## Definie Button Actions KennwerteBau
        self.ui.cmdStartQueryKenn.clicked.connect(self.start_query)
        self.ui.cmdCalcKenn.clicked.connect(self.mw_calc_kenn)
        self.ui.cmdKennDetailBack.clicked.connect(self.show_kennwerte_bau_main)
        self.ui.cmdKennDetailLVBack.clicked.connect(self.show_kennwerte_bau_detail)

        ## Definie Button Actions Verwaltung
        self.ui.cmdVerwaltungIndex.clicked.connect(self.show_verwaltung_index)
        self.ui.cmdVerwaltungProjekt.clicked.connect(self.show_verwaltung_projekt)
        self.ui.cmdIndexSave.clicked.connect(self.index_execute)

    def closeEvent(self, a0: QtGui.QCloseEvent) -> None:
        pass

    ## Allgemein
    def start_query(self):
        if self.ui.stackedMain.currentIndex() == 0:
            self.db_fill_table()
        elif self.ui.stackedMain.currentIndex() == 1:
            self.db_fill_table_kenn()

    ## Window change functions
    def open_index(self):
        self.partnerDialogCalc.show()

    def show_lv(self):
        self.ui.stackedMain.setCurrentWidget(self.ui.pageLVSuche)
        self.show_lv_main()

    def show_lv_main(self):
        self.ui.stackedLVPos.setCurrentWidget(self.ui.pageMainTable)

    def show_kennwerte_bau(self):
        self.ui.stackedMain.setCurrentWidget(self.ui.pageKennwerteBau)

    def show_kennwerte_bau_main(self):
        self.ui.stackedKenn.setCurrentWidget(self.ui.pageMainTableKenn)
        self.show_kennwerte_bau()

    def show_kennwerte_bau_detail(self):
        self.ui.stackedKenn.setCurrentWidget(self.ui.pageDetailKenn)
        self.show_kennwerte_bau()

    def show_kennwerte_bau_lv_detail(self):
        self.ui.stackedKenn.setCurrentWidget(self.ui.pageDetailLVKenn)
        self.show_kennwerte_bau()

    def show_verwaltung(self):
        self.ui.stackedMain.setCurrentWidget(self.ui.pageVerwaltung)
        self.show_verwaltung_index()

    def show_verwaltung_index(self):
        self.ui.stackedVerwaltung.setCurrentWidget(self.ui.pageVerwaltungIndex)

    def show_verwaltung_projekt(self):
        self.ui.stackedVerwaltung.setCurrentWidget(self.ui.pageVerwaltungProjekt)

    ## LV-Position
    def set_controls_lv(self):
        ## Anlegen des Query-Objekts
        self.query = QtSql.QSqlQuery()

        ## Abfrage Kategorie
        self.qry_kat = """
            SELECT DISTINCT [Gewerke Kat.]
            FROM GewerkeView
            ORDER BY [Gewerke Kat.] ASC
        """
        self.query.exec(self.qry_kat)

        self.ui.cboKategorie.addItem("-")
        while self.query.next():
            if self.query.value(0) == "Architektur":
                self.ui.cboKategorie.addItem("Baugewerk")
            else:
                self.ui.cboKategorie.addItem(self.query.value(0))

        ## Abfrage Einheiten
        self.qry_eh = """
            SELECT DISTINCT Einheit
            FROM GewerkeView
            ORDER BY Einheit ASC
        """
        self.query.exec(self.qry_eh)

        self.ui.cboEinheit.addItem("-")
        while self.query.next():
            self.ui.cboEinheit.addItem(self.query.value(0))

        ##Abfrage Preisbasis
        self.qry_pb = """
            SELECT  MIN(Preisbasis), MAX(Preisbasis)
            FROM GewerkeView
        """
        self.query.exec(self.qry_pb)

        self.query.first()
        low = self.query.value(0).toPyDateTime()
        high = self.query.value(1).toPyDateTime()

        self.ui.datPBVon.setDateRange(QDate(low.year, low.month, low.day), QDate(high.year, high.month, high.day))
        self.ui.datPBBis.setDateRange(QDate(low.year, low.month, low.day), QDate(high.year, high.month, high.day))
        self.ui.datPBVon.setDate(QDate(low.year, low.month, low.day))
        self.ui.datPBBis.setDate(QDate(high.year, high.month, high.day))

        ## Abfrage Menge
        self.qry_menge = """
            SELECT FORMAT(MIN([Menge]), 'N', 'de-de'), FORMAT(MAX([Menge]), 'N', 'de-de')
            FROM GewerkeView
        """
        self.query.exec(self.qry_menge)

        self.query.first()
        low = string_to_float(self.query.value(0), 2)
        high = string_to_float(self.query.value(1), 2)

        self.ui.spnMengeMin.setRange(low, high)
        self.ui.spnMengeMax.setRange(low, high)
        self.ui.spnMengeMin.setValue(low)
        self.ui.spnMengeMax.setValue(high)

        ## Abfrage Projektliste
        self.qry_projekt = """
            SELECT  GZ, Projektname
            FROM ProjekteView
        """
        self.query.exec(self.qry_projekt)

        # self.ui.lstProjekte.addItem("-")
        while self.query.next():
            self.ui.lstProjekte.addItem(self.query.value(0) + " / " + self.query.value(1))

    def reset_controls_lv(self):
        ## Set standard-values to control elements
        self.setCursor(QCursor(QtCore.Qt.WaitCursor))

        self.ui.cboKategorie.clear()
        self.ui.cboEinheit.clear()
        self.ui.lstProjekte.clear()
        self.ui.txtSearchLong.setText("")
        self.ui.txtSearchShort.setText("")
        self.ui.txtSearchLG.setText("")
        self.ui.txtSearchGewerk.setText("")
        self.ui.txtSearchLVPos.setText("")
        self.set_controls_lv()

        self.setCursor(QCursor(QtCore.Qt.ArrowCursor))

    def generate_menu_lv(self):
        # bar = self.parent.menuBar()
        top_menu = QMenu(self)

        menu = top_menu.addMenu("Menu")
        # config = menu.addMenu("Configuration ...")

        details = menu.addAction("Details")
        index = menu.addAction("Export Indexrechner")

        menu.addSeparator()

        error = menu.addAction("Fehler in LVPos")

        # ehp_calc = menu.addMenu("EHP-Kalkulation")
        # ehp_mittel = ehp_calc.addAction("Mittelwert")
        # ehp_3best = ehp_calc.addAction("3 Bestbieter")
        # ehp_billig = ehp_calc.addAction("Billigstbieter")
        # ehp_median = ehp_calc.addAction("Median")

        #
        # config1 = config.addAction("Config1")
        # config2 = config.addAction("Config2")
        # config3 = config.addAction("Config3")

        action = menu.exec_(QtGui.QCursor.pos())

        if action == details:
            self.show_lv_details()
        elif action == index:
            self.partnerDialogCalc.show()
            self.partnerDialogCalc.import_preis()
        elif action == error:
            self.send_wrong_pos()
        # elif action == ehp_mittel:
        #     if self.ui.grdAbfrage.rowCount() == 0:
        #         return
        #     else:
        #         self.export_ehp("Mittel")
        #     pass
        # elif action == ehp_3best:
        #     if self.ui.grdAbfrage.rowCount() == 0:
        #         return
        #     else:
        #         self.export_ehp("3Best")
        #     pass
        # elif action == ehp_billig:
        #     if self.ui.grdAbfrage.rowCount() == 0:
        #         return
        #     else:
        #         self.export_ehp("Billig")
        #     pass
        # elif action == ehp_median:
        #     if self.ui.grdAbfrage.rowCount() == 0:
        #         return
        #     else:
        #         self.export_ehp("Median")
        #     pass

    def sql_main_lv(self):
        sql_main = """
            SELECT DISTINCT LVPos,
                            [LVPos Stichwort],
                            FORMAT(Menge, 'N', 'de-de'),
                            Einheit,
                            FORMAT([Preis Mittelwert VAL], 'N', 'de-de'),
                            FORMAT([Preis 3 Bestbieter VAL], 'N', 'de-de'),
                            FORMAT([Preis Billigstbieter VAL], 'N', 'de-de'),
                            FORMAT([Preis Median VAL], 'N', 'de-de'),
                            FORMAT(Preisbasis, 'yyy-MM-dd'),
                            GZ,
                            Projektname,
                            Gewerkebezeichnung,
                            [LVPos Z-Position],
                            HG,
                            OG,
                            FORMAT([Preis Mittelwert], 'N', 'de-de'),
                            FORMAT([Preis 3 Bestbieter], 'N', 'de-de'),
                            FORMAT([Preis Billigstbieter], 'N', 'de-de'),
                            FORMAT([Preis Median], 'N', 'de-de')
            FROM GewerkeViewOhneBieter
        """

        ## Textsuche kurz
        sql_short = ""
        txt_short_zaehler = 1
        txt_short = self.ui.txtSearchShort.text()
        if '++' in txt_short:
            txt_short_split = txt_short.split("++")
            txt_short_anzahl = len(txt_short_split)
            for word in txt_short_split:
                if txt_short_zaehler < txt_short_anzahl:
                    sql_short = sql_short + "[LVPos Stichwort + Langtext] LIKE '%" + word + "%' AND "
                else:
                    sql_short = sql_short + "[LVPos Stichwort + Langtext] LIKE '%" + word + "%'"
                txt_short_zaehler = txt_short_zaehler + 1
        elif '--' in txt_short:
            txt_short_split = txt_short.split("--")
            txt_short_anzahl = len(txt_short_split)
            for word in txt_short_split:
                if txt_short_zaehler < txt_short_anzahl:
                    sql_short = sql_short + "[LVPos Stichwort + Langtext] LIKE '%" + word + "%' OR "
                else:
                    sql_short = sql_short + "[LVPos Stichwort + Langtext] LIKE '%" + word + "%'"
                txt_short_zaehler = txt_short_zaehler + 1
        else:
            sql_short = sql_short + "[LVPos Stichwort + Langtext] LIKE '%" + txt_short + "%'"

        ## Textsuche lang
        sql_long = ""
        txt_long_zaehler = 1
        txt_long = self.ui.txtSearchLong.text()
        if '++' in txt_long:
            txt_long_split = txt_long.split("++")
            txt_long_anzahl = len(txt_long_split)
            for word in txt_long_split:
                if txt_long_zaehler < txt_long_anzahl:
                    sql_long = sql_long + "[LVPos Stichwort + Langtext + ULG Langtext] LIKE '%" + word + "%' AND "
                else:
                    sql_long = sql_long + "[LVPos Stichwort + Langtext + ULG Langtext] LIKE '%" + word + "%'"
                txt_long_zaehler = txt_long_zaehler + 1
        elif '--' in txt_long:
            txt_long_split = txt_long.split("--")
            txt_long_anzahl = len(txt_long_split)
            for word in txt_long_split:
                if txt_long_zaehler < txt_long_anzahl:
                    sql_long = sql_long + "[LVPos Stichwort + Langtext + ULG Langtext] LIKE '%" + word + "%' OR "
                else:
                    sql_long = sql_long + "[LVPos Stichwort + Langtext + ULG Langtext] LIKE '%" + word + "%'"
                txt_long_zaehler = txt_long_zaehler + 1
        else:
            sql_long = sql_long + "[LVPos Stichwort + Langtext + ULG Langtext] LIKE '%" + txt_long + "%'"

        ## Gewerkesuche
        sql_gew = ""
        txt_gew_zaehler = 1
        txt_gew = self.ui.txtSearchGewerk.text()
        if '++' in txt_gew:
            txt_gew_split = txt_gew.split("++")
            txt_gew_anzahl = len(txt_gew_split)
            for word in txt_gew_split:
                if txt_gew_zaehler < txt_gew_anzahl:
                    sql_gew = sql_gew + "[Gewerkebezeichnung] LIKE '%" + word + "%' AND "
                else:
                    sql_gew = sql_gew + "[Gewerkebezeichnung] LIKE '%" + word + "%'"
                txt_gew_zaehler = txt_gew_zaehler + 1
        elif '--' in txt_gew:
            txt_gew_split = txt_gew.split("--")
            txt_gew_anzahl = len(txt_gew_split)
            for word in txt_gew_split:
                if txt_gew_zaehler < txt_gew_anzahl:
                    sql_gew = sql_gew + "[Gewerkebezeichnung] LIKE '%" + word + "%' OR "
                else:
                    sql_gew = sql_gew + "[Gewerkebezeichnung] LIKE '%" + word + "%'"
                txt_gew_zaehler = txt_gew_zaehler + 1
        else:
            sql_gew = sql_gew + "[Gewerkebezeichnung] LIKE '%" + txt_gew + "%'"

        ## LG-Suche
        sql_lg = ""
        txt_lg_zaehler = 1
        txt_lg = self.ui.txtSearchLG.text()
        if '++' in txt_lg:
            txt_lg_split = txt_lg.split("++")
            txt_lg_anzahl = len(txt_lg_split)
            for word in txt_lg_split:
                if txt_lg_zaehler < txt_lg_anzahl:
                    sql_lg = sql_lg + "[LG Stichwort] LIKE '%" + word + "%' AND "
                else:
                    sql_lg = sql_lg + "[LG Stichwort] LIKE '%" + word + "%'"
                txt_lg_zaehler = txt_lg_zaehler + 1
        elif '--' in txt_lg:
            txt_lg_split = txt_lg.split("--")
            txt_lg_anzahl = len(txt_lg_split)
            for word in txt_lg_split:
                if txt_lg_zaehler < txt_lg_anzahl:
                    sql_lg = sql_lg + "[LG Stichwort] LIKE '%" + word + "%' OR "
                else:
                    sql_lg = sql_lg + "[LG Stichwort] LIKE '%" + word + "%'"
                txt_lg_zaehler = txt_lg_zaehler + 1
        else:
            sql_lg = sql_lg + "[LG Stichwort] LIKE '%" + txt_lg + "%'"

        ## Kategorie
        if self.ui.cboKategorie.currentText() == "-":
            txt_kat = ""
        elif self.ui.cboKategorie.currentText() == "Baugewerk":
            txt_kat = "Architektur"
        else:
            txt_kat = self.ui.cboKategorie.currentText()
        sql_kat = "[Gewerke Kat.] LIKE '%" + txt_kat + "%'"

        ## Preisbasis von
        txt_start = self.ui.datPBVon.text()
        start = datetime.strptime(txt_start, '%d.%m.%Y')
        sql_start = "Preisbasis >= '" + str(start.date()) + "'"

        ## Preisbasis bis
        txt_ende = self.ui.datPBBis.text()
        ende = datetime.strptime(txt_ende, '%d.%m.%Y')
        sql_ende = "Preisbasis <= '" + str(ende.date()) + "'"

        ## Menge min
        txt_min = self.ui.spnMengeMin.value()
        sql_min = "Menge >= " + str(txt_min)

        ## Menge max
        txt_max = self.ui.spnMengeMax.value()
        sql_max = "Menge <= " + str(txt_max)

        ## LV Pos.
        if self.ui.txtSearchLVPos.text() == "":
            txt_LVPos = "%%"
        else:
            txt_LVPos = self.ui.txtSearchLVPos.text()
        sql_LVPos = "[LVPos] LIKE '" + txt_LVPos + "'"

        ## Einheit
        if self.ui.cboEinheit.currentText() == "-":
            txt_EH = "%%"
        else:
            txt_EH = self.ui.cboEinheit.currentText()
        sql_EH = "[Einheit] LIKE '" + str(txt_EH).strip() + "'"

        ## Projekte
        projects = self.ui.lstProjekte.selectedItems()
        sql_proj = ""
        txt_proj_anzahl = len(projects)
        txt_proj_zaehler = 1
        if txt_proj_anzahl == 0:
            sql_proj = "[GZ] LIKE '%%'"
        else:
            for project in projects:
                project_split = project.text().split("/")
                project_final = project_split[0].strip()
                if txt_proj_zaehler < txt_proj_anzahl:
                    sql_proj = sql_proj + "[GZ] LIKE '%" + project_final + "%' OR "
                else:
                    sql_proj = sql_proj + "[GZ] LIKE '%" + project_final + "%'"
                txt_proj_zaehler = txt_proj_zaehler + 1

        ## WHERE
        sql_where = "WHERE \n"

        ## SQL Main
        sql = sql_main + sql_where + \
              "(" + sql_short + ") AND \n" + \
              "(" + sql_long + ") AND \n" + \
              "(" + sql_gew + ") AND \n" + \
              "(" + sql_lg + ") AND \n" + \
              sql_LVPos + " AND \n" + \
              sql_kat + " AND \n" + \
              sql_start + " AND \n" + \
              sql_ende + " AND \n" + \
              sql_min + " AND \n" + \
              sql_max + " AND \n" + \
              "(" + sql_proj + ") AND \n" + \
              sql_EH

        ## Count Rows
        sql_count = """
            SELECT COUNT(DISTINCT(CONCAT(LVPos,
                            GZ,
                            Gewerkebezeichnung,
                            [LVPos Z-Position],
                            HG,
                            OG)))
            FROM GewerkeView
        """

        sql_count_final = sql_count + sql_where + \
                          "(" + sql_short + ") AND \n" + \
                          "(" + sql_long + ") AND \n" + \
                          "(" + sql_gew + ") AND \n" + \
                          "(" + sql_lg + ") AND \n" + \
                          sql_LVPos + " AND \n" + \
                          sql_kat + " AND \n" + \
                          sql_start + " AND \n" + \
                          sql_ende + " AND \n" + \
                          sql_min + " AND \n" + \
                          sql_max + " AND \n" + \
                          "(" + sql_proj + ") AND \n" + \
                          sql_EH

        return sql, sql_count_final

    def db_fill_table(self):
        self.setCursor(QCursor(QtCore.Qt.WaitCursor))
        try:
            start_time = datetime.now()

            sql = self.sql_main_lv()
            # print(sql[1])

            # ## Positionsanzahl
            # self.qry1 = sql[1]
            # self.query = QtSql.QSqlQuery()
            # self.query.exec(self.qry1)
            # self.query.first()
            # self.string = format_decimal(self.query.value(0), format='#,##0', locale='de_DE')

            model = MySqlModel()
            model.setQuery(sql[0])
            self.string = str(model.rowCount())

            col = 0
            headers = ['LVPos', 'LVPos Stichwort', 'Menge', 'EH', 'EHP Mittelwert Val', 'EHP 3 Bestbieter VAL', 'EHP Biligstbieter VAL', 'EHP Median VAL',
                       'Preisbasis', 'GZ', 'Projektname', 'Gewerkebezeichnung',
                       'Z-Position', 'HG', 'OG', 'EHP Mittelwert', 'EHP 3 bestbieter', 'EHP Billigstbieter', 'EHP Median']
            for header in headers:
                model.setHeaderData(col, QtCore.Qt.Horizontal, header)
                col += 1

            # print((datetime.now() - start_time).total_seconds())

            view = self.ui.grdView
            view.setModel(model)
            view.resizeColumnToContents(0)
            view.setColumnWidth(1, 300)
            # view.resizeColumnToContents(1)
            view.resizeColumnToContents(2)
            view.resizeColumnToContents(3)
            view.setColumnWidth(4, 100)
            view.setColumnWidth(5, 100)
            view.setColumnWidth(6, 100)
            view.setColumnWidth(7, 100)
            view.resizeColumnToContents(8)
            view.resizeColumnToContents(9)
            view.resizeColumnToContents(10)
            view.resizeColumnToContents(11)
            # view.setColumnWidth(0, 200)
            # view.horizontalHeader().setStretchLastSection(True)
            # view.resizeColumnsToContents()
            self.check_screen = CheckScreen()

            end_time = datetime.now()
            diff = (end_time - start_time).total_seconds()
            if model.canFetchMore() == True:
                self.ui.lblDataCount.setText("Abfragedauer: " + self.string + "+ Positionen in " + str(round(diff, 3)) + "s")
            else:
                self.ui.lblDataCount.setText("Abfragedauer: " + self.string + " Positionen in " + str(round(diff, 3)) + "s")

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Fehler bei SQL!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        self.setCursor(QCursor(QtCore.Qt.ArrowCursor))

    def mw_calc_lv(self):
        list_mittelwert = []
        list_best3 = []
        list_min = []
        list_median = []

        try:
            rows = self.ui.grdView.selectionModel().selectedRows()
            for row in rows:
                for col in range(4, 8):
                    if col == 4:
                        value = self.ui.grdView.model().index(row.row(), col)
                        data_value = self.ui.grdView.model().data(value)
                        list_mittelwert.append(string_to_float(data_value, 2))
                    elif col == 5:
                        value = self.ui.grdView.model().index(row.row(), col)
                        data_value = self.ui.grdView.model().data(value)
                        list_best3.append(string_to_float(data_value, 2))
                    elif col == 6:
                        value = self.ui.grdView.model().index(row.row(), col)
                        data_value = self.ui.grdView.model().data(value)
                        list_min.append(string_to_float(data_value, 2))
                    elif col == 7:
                        value = self.ui.grdView.model().index(row.row(), col)
                        data_value = self.ui.grdView.model().data(value)
                        list_median.append(string_to_float(data_value, 2))

            self.ui.lblMWMittelwert.setText(float_to_string(statistics.mean(list_mittelwert)))
            self.ui.lblMW3Best.setText(float_to_string(statistics.mean(list_best3)))
            self.ui.lblMWBilligst.setText(float_to_string(statistics.mean(list_min)))
            self.ui.lblMWMedian.setText(float_to_string(statistics.mean(list_median)))

        except:
            self.ui.lblMWMittelwert.setText(float_to_string(statistics.mean(0)))
            self.ui.lblMW3Best.setText(float_to_string(statistics.mean(0)))
            self.ui.lblMWBilligst.setText(float_to_string(statistics.mean(0)))
            self.ui.lblMWMedian.setText(float_to_string(statistics.mean(0)))

    def export_excel_selected(self):
        if self.ui.stackedMain.currentIndex() == 0:
            try:
                rows = self.ui.grdView.selectionModel().selectedRows()
                if len(rows) == 0:
                    msg = QMessageBox()
                    msg.setWindowIcon(QtGui.QIcon(company_icon))
                    msg.setIcon(QMessageBox.Warning)
                    msg.setText("Kein Position gewählt!")
                    msg.setWindowTitle("Preisdatenbank")
                    msg.setStandardButtons(QMessageBox.Ok)
                    msg.exec_()
                    return

                file_target = QFileDialog.getSaveFileName(self, "Speichern unter", r"\\firma.local\dfs\Projekte",
                                                          "Excel-Dateien (*.xlsx)")
                if file_target[0] == '':
                    msg = QMessageBox()
                    msg.setWindowIcon(QtGui.QIcon(company_icon))
                    msg.setIcon(QMessageBox.Warning)
                    msg.setText("Es wurde kein Speicherpfad gewählt!")
                    msg.setWindowTitle("Preisdatenbank")
                    msg.setStandardButtons(QMessageBox.Ok)
                    msg.exec_()
                    return

                workbook_base = load_workbook(filename=path_templates_LV, read_only=False, keep_vba=False)
                base = workbook_base.worksheets[0]
                xl_row = 2
                for row in rows:
                    for col in range(0, 19):
                        if col >= 0 and col <= 1 or col == 3 or col >= 8 and col <= 14:
                            value = self.ui.grdView.model().index(row.row(), col)
                            data_value = self.ui.grdView.model().data(value)
                            base.cell(row=xl_row, column=col + 1).value = data_value

                        if col == 2 or col >= 4 and col <= 7 or col >= 15 and col <= 18:
                            value = self.ui.grdView.model().index(row.row(), col)
                            data_value = self.ui.grdView.model().data(value)
                            base.cell(row=xl_row, column=col + 1).value = string_to_float(data_value, 2)

                    xl_row += 1
                workbook_base.save(file_target[0])

                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Information)
                msg.setText("LV-Positionen erfolgreich exportiert!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

            except:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Keine Positionen vorhanden!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

        elif self.ui.stackedMain.currentIndex() == 1:
            try:
                rows = self.ui.grdViewKenn.selectionModel().selectedRows()
                if len(rows) == 0:
                    msg = QMessageBox()
                    msg.setWindowIcon(QtGui.QIcon(company_icon))
                    msg.setIcon(QMessageBox.Warning)
                    msg.setText("Kein Position gewählt!")
                    msg.setWindowTitle("Preisdatenbank")
                    msg.setStandardButtons(QMessageBox.Ok)
                    msg.exec_()
                    return

                file_target = QFileDialog.getSaveFileName(self, "Speichern unter", r"\\firma.local\dfs\Projekte",
                                                          "Excel-Dateien (*.xlsx)")
                if file_target[0] == '':
                    msg = QMessageBox()
                    msg.setWindowIcon(QtGui.QIcon(company_icon))
                    msg.setIcon(QMessageBox.Warning)
                    msg.setText("Es wurde kein Speicherpfad gewählt!")
                    msg.setWindowTitle("Preisdatenbank")
                    msg.setStandardButtons(QMessageBox.Ok)
                    msg.exec_()
                    return

                workbook_base = load_workbook(filename=path_templates_Kennwerte, read_only=False, keep_vba=False)
                base = workbook_base.worksheets[0]
                xl_row = 2
                for row in rows:
                    for col in range(0, 19):
                        if col >= 0 and col <= 6:
                            value = self.ui.grdViewKenn.model().index(row.row(), col)
                            data_value = self.ui.grdViewKenn.model().data(value)
                            base.cell(row=xl_row, column=col + 1).value = data_value

                        if col >= 7 and col <= 11:
                            value = self.ui.grdViewKenn.model().index(row.row(), col)
                            data_value = self.ui.grdViewKenn.model().data(value)
                            base.cell(row=xl_row, column=col + 1).value = string_to_float(data_value, 2)

                    xl_row += 1
                workbook_base.save(file_target[0])

                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Information)
                msg.setText("Kennwert-Positionen erfolgreich exportiert!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

            except:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Keine Positionen vorhanden!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

    def show_lv_details(self):
        try:
            rows = self.ui.grdView.selectionModel().selectedRows()
            if len(rows) == 0:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Kein Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) > 1:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Mehr als eine Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) == 1:
                self.ui.stackedLVPos.setCurrentWidget(self.ui.pageDetail)
                list_values = []
                for col in range(0, 19):
                    # 0, 1, 2, 3, 8, 9, 11, 12, 13, 14
                    if col >= 0 and col <= 3 or col >= 8 and col <= 9 or col >= 11 and col <= 14:
                        value = self.ui.grdView.model().index(rows[0].row(), col)
                        data_value = self.ui.grdView.model().data(value)
                        # date_fin = datetime.strptime(data_value, '%Y-%m-%d')
                        list_values.append(data_value)

                gz = list_values[5]
                gewerkebezeichnung = list_values[6]
                preisbasis = list_values[4]
                hg = list_values[8]
                og = list_values[9]
                lvpos = list_values[0]
                lvzpos = list_values[7]
                lv_stichwort = list_values[1]
                menge = list_values[2]
                eh = list_values[3]

                lvpos_final = lvpos + lvzpos

                self.ui.lblDetailLVPos.setText(lvpos_final)
                self.ui.lblDetailStichwort.setText(lv_stichwort)
                self.ui.lblDetailPB.setText(preisbasis)
                self.ui.lblDetailMenge.setText(menge)
                self.ui.lblDetailEinheit.setText(eh)

                model = MySqlModel2()
                sql_bieter = """
                            SELECT DISTINCT Bieter, FORMAT([Preis], 'N', 'de-de')
                            FROM GewerkeView
                            WHERE GZ LIKE '""" + gz + """' AND
                                    Gewerkebezeichnung LIKE '""" + gewerkebezeichnung + """' AND
                                    (HG LIKE '""" + hg + """' OR HG is NULL) AND
                                    (OG LIKE '""" + og + """' OR OG is NULL) AND
                                    LVPos LIKE '""" + lvpos + """' AND
                                    ([LVPos Z-Position] LIKE '""" + lvzpos + """' OR [LVPos Z-Position] is NULL)
                        """
                model.setQuery(sql_bieter)

                col = 0
                headers = ['Bieter', 'EHP']
                for header in headers:
                    model.setHeaderData(col, QtCore.Qt.Horizontal, header)
                    col += 1

                view = self.ui.grdBieter
                view.setModel(model)
                view.setColumnWidth(0, 100)

                sql_langtext = """
                    SELECT DISTINCT [LVPos Stichwort], [LVPos Langtext], [ULG Stichwort], [UnterleistungsgruppeLangtext], [LG Stichwort], [LG Langtext]
                    FROM GewerkeView
                    WHERE GZ LIKE '""" + gz + """' AND
                            Gewerkebezeichnung LIKE '""" + gewerkebezeichnung + """' AND
                            (HG LIKE '""" + hg + """' OR HG is NULL) AND
                            (OG LIKE '""" + og + """' OR OG is NULL) AND
                            LVPos LIKE '""" + lvpos + """' AND
                            ([LVPos Z-Position] LIKE '""" + lvzpos + """' OR [LVPos Z-Position] is NULL)
                """

                self.query = QtSql.QSqlQuery()
                self.query.exec(sql_langtext)
                self.query.first()

                lvpos_stichwort = self.query.value(0)
                lvpos_langtext = self.query.value(1)
                lvpos_komplett = lvpos_stichwort + "\n\n" + lvpos_langtext
                self.ui.lblLangtextLVPos.setText(lvpos_komplett)

                ulg_stichwort = self.query.value(2)
                ulg_langtext = self.query.value(3)
                ulg_komplett = ulg_stichwort + "\n\n" + ulg_langtext
                self.ui.lblLangtextULG.setText(ulg_komplett)

                lg_stichwort = self.query.value(4)
                lg_langtext = self.query.value(5)
                lg_komplett = lg_stichwort + "\n\n" + lg_langtext
                self.ui.lblLangtextLG.setText(lg_komplett)

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Keine Positionen vorhanden!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def send_wrong_pos(self):
        try:
            rows = self.ui.grdView.selectionModel().selectedRows()
            if len(rows) == 0:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Kein Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) > 1:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Mehr als eine Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) == 1:
                heute = datetime.now()
                zeitpunkt = heute.strftime('%Y%m%d_%H%M')

                filename_log = zeitpunkt + "_Falsche_Position.log"
                filepath_log = r"\\firma.local\dfs\Laufwerk-Z\Daten - Intern\Kosten\09 SQL Preisdatenbank\07_Python Oberfläche\Logging" + "\\" + filename_log
                logging.basicConfig(filename=filepath_log, encoding='utf-8', level=logging.DEBUG)
                user = getpass.getuser().capitalize()

                list_values = []
                for col in range(0, 19):
                    # 0, 1, 2, 3, 8, 9, 11, 12, 13, 14
                    if col >= 0 and col <= 3 or col >= 8 and col <= 9 or col >= 11 and col <= 14:
                        value = self.ui.grdView.model().index(rows[0].row(), col)
                        data_value = self.ui.grdView.model().data(value)
                        # date_fin = datetime.strptime(data_value, '%Y-%m-%d')
                        list_values.append(data_value)

                gz = list_values[5]
                gewerkebezeichnung = list_values[6]
                preisbasis = list_values[4]
                hg = list_values[8]
                og = list_values[9]
                lvpos = list_values[0]
                lvzpos = list_values[7]
                lv_stichwort = list_values[1]
                menge = list_values[2]
                eh = list_values[3]

                logging.info("User: " + user)
                logging.warning("LV-Position: " + lvpos)
                logging.warning("Z-Position: " + lvzpos)
                logging.warning("Stichwort: " + lv_stichwort)
                logging.warning("Menge: " + menge)
                logging.warning("Einheit: " + eh)
                logging.warning("HG: " + hg)
                logging.warning("OG: " + og)
                logging.warning("GZ: " + gz)
                logging.warning("Gewerkebezeichnung: " + gewerkebezeichnung)
                logging.warning("Preisbasis: " + preisbasis)

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Keine Positionen vorhanden!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    ## Kennwerte Bau
    kennwerteStatusUmbauNeubau = ""

    def set_controls_kenn(self):
        ## Anlegen des Query-Objekts
        self.query = QtSql.QSqlQuery()

        ## Abfrage Projektliste
        self.qry_projekt = """
            SELECT DISTINCT GZ,Projektname
            FROM KennwerteProjekteFinalView
            ORDER BY [GZ] ASC
        """
        self.query.exec(self.qry_projekt)

        # self.ui.lstProjekte.addItem("-")
        while self.query.next():
            self.ui.lstProjekteKenn.addItem(self.query.value(0) + " / " + self.query.value(1))

        ## Abfrage Code
        self.qry_code = """
            SELECT DISTINCT [Code], [Bezeichnung]
            FROM KennwerteProjekteFinalView
            ORDER BY [Code] ASC
        """
        self.query.exec(self.qry_code)

        while self.query.next():
            self.ui.lstCode.addItem(self.query.value(0) + " - " + self.query.value(1))

        ## Abfrage Gebäudeart
        self.qry_kat = """
            SELECT DISTINCT [Objektart]
            FROM KennwerteProjekteFinalView
            ORDER BY [Objektart] ASC
        """
        self.query.exec(self.qry_kat)

        self.ui.cboArt.addItem("-")
        while self.query.next():
            self.ui.cboArt.addItem(self.query.value(0))

        ## Abfrage Menge
        self.qry_menge = """
            SELECT FORMAT(MIN(MengeFinal), 'N', 'de-de'), FORMAT(MAX(MengeFinal), 'N', 'de-de')
            FROM KennwerteProjekteFinalView
        """
        self.query.exec(self.qry_menge)

        self.query.first()
        low = string_to_float(self.query.value(0), 2)
        high = string_to_float(self.query.value(1), 2)

        self.ui.spnBGFMin.setRange(low, high)
        self.ui.spnBGFMax.setRange(low, high)
        self.ui.spnBGFMin.setValue(low)
        self.ui.spnBGFMax.setValue(high)

        ## Umbaustatus
        self.ui.cboUmbauStatus.addItem("-")
        self.ui.cboUmbauStatus.addItem("Neubau")
        self.ui.cboUmbauStatus.addItem("Umbau")

    def reset_controls_kenn(self):
        ## Set standard-values to control elements
        self.setCursor(QCursor(QtCore.Qt.WaitCursor))

        self.ui.txtSearchCodeText.setText("")
        self.ui.cboArt.clear()
        self.ui.cboUmbauStatus.clear()
        self.ui.lstCode.clear()
        self.ui.lstProjekteKenn.clear()
        self.set_controls_kenn()

        self.setCursor(QCursor(QtCore.Qt.ArrowCursor))

    def generate_menu_kenn(self):
        # bar = self.parent.menuBar()
        top_menu = QMenu(self)

        menu = top_menu.addMenu("Menu")

        details = menu.addAction("Details")

        action = menu.exec_(QtGui.QCursor.pos())

        if action == details:
            self.show_kenn_details()

    def generate_menu_kenn_detail(self):
        # bar = self.parent.menuBar()
        top_menu = QMenu(self)

        menu = top_menu.addMenu("Menu")

        details = menu.addAction("Details")

        action = menu.exec_(QtGui.QCursor.pos())

        if action == details:
            self.show_kenn_details_lvpos()

    def sql_main_kenn(self):
        global kennwerteStatusUmbauNeubau
        kennwerteStatusUmbauNeubau = self.ui.cboUmbauStatus.currentText()

        if kennwerteStatusUmbauNeubau == "Neubau":
            sql_main = """SELECT DISTINCT GZ, Projektname, Objektart, Code, Bezeichnung, Einheit, Bezug, FORMAT(MengeFinal, 'N', 'de-de'), FORMAT([Kennwert Mittelwert], 'N', 'de-de'), FORMAT([Kennwert Mittelwert 3 Bestbieter], 'N', 'de-de'), FORMAT([Kennwert Billigstbieter], 'N', 'de-de'), FORMAT([Kennwert Median], 'N', 'de-de'), 'Neubau' AS Status
                FROM KennwerteProjekteFinalView
                """
        elif kennwerteStatusUmbauNeubau == "Umbau":
            sql_main = """SELECT DISTINCT GZ, Projektname, Objektart, Code, Bezeichnung, Einheit, Bezug, FORMAT(MengeFinalUmbau, 'N', 'de-de'), FORMAT([Kennwert Mittelwert Umbau], 'N', 'de-de'), FORMAT([Kennwert Mittelwert 3 Bestbieter Umbau], 'N', 'de-de'), FORMAT([Kennwert Billigstbieter Umbau], 'N', 'de-de'), FORMAT([Kennwert Median Umbau], 'N', 'de-de'), 'Umbau' AS Status
                            FROM KennwerteProjekteFinalView
                            """
        else:
            sql_main = """SELECT GZ,Projektname,Objektart,Code,Bezeichnung,Einheit,Bezug,FORMAT(Menge, 'N', 'de-de'), FORMAT(Mittelwert, 'N', 'de-de'), FORMAT(Mittelwert3Bestbieter, 'N', 'de-de'), FORMAT(Billigstbieter, 'N', 'de-de'), FORMAT(Median, 'N', 'de-de'),Status
                          FROM KennwerteProjekteFinalViewMitStatus
                          """

        ## Codesuche
        sql_codeBez = ""
        txt_codeBez_zaehler = 1
        txt_codeBez = self.ui.txtSearchCodeText.text()
        if '++' in txt_codeBez:
            txt_codeBez_split = txt_codeBez.split("++")
            txt_codeBez_anzahl = len(txt_codeBez_split)
            for word in txt_codeBez_split:
                word = word.strip()
                if txt_codeBez_zaehler < txt_codeBez_anzahl:
                    sql_codeBez = sql_codeBez + "Bezeichnung LIKE '%" + word + "%' AND "
                else:
                    sql_codeBez = sql_codeBez + "Bezeichnung LIKE '%" + word + "%'"
                txt_codeBez_zaehler = txt_codeBez_zaehler + 1
        elif '--' in txt_codeBez:
            txt_codeBez_split = txt_codeBez.split("--")
            txt_codeBez_anzahl = len(txt_codeBez_split)
            for word in txt_codeBez_split:
                word = word.strip()
                if txt_codeBez_zaehler < txt_codeBez_anzahl:
                    sql_codeBez = sql_codeBez + "Bezeichnung LIKE '%" + word + "%' OR "
                else:
                    sql_codeBez = sql_codeBez + "Bezeichnung LIKE '%" + word + "%'"
                txt_codeBez_zaehler = txt_codeBez_zaehler + 1
        else:
            sql_codeBez = sql_codeBez + "Bezeichnung LIKE '%" + txt_codeBez + "%'"

        ## Gebäudeart
        if self.ui.cboArt.currentText() == "-":
            txt_art = ""
        else:
            txt_art = self.ui.cboArt.currentText()
        sql_art = "Objektart LIKE '%" + txt_art + "%'"

        ## BGF min
        if kennwerteStatusUmbauNeubau == "Neubau":
            txt_min = self.ui.spnBGFMin.value()
            sql_min = "MengeFinal >= " + str(txt_min)
        elif kennwerteStatusUmbauNeubau == "Umbau":
            txt_min = self.ui.spnBGFMin.value()
            sql_min = "MengeFinalUmbau >= " + str(txt_min)
        else:
            txt_min = self.ui.spnBGFMin.value()
            sql_min = "Menge >= " + str(txt_min)
        ## BGF max
        if kennwerteStatusUmbauNeubau == "Neubau":
            txt_max = self.ui.spnBGFMax.value()
            sql_max = "MengeFinal <= " + str(txt_max)
        elif kennwerteStatusUmbauNeubau == "Umbau":
            txt_max = self.ui.spnBGFMax.value()
            sql_max = "MengeFinalUmbau <= " + str(txt_max)
        else:
            txt_max = self.ui.spnBGFMax.value()
            sql_max = "Menge <= " + str(txt_max)

        ## Projekte
        projects = self.ui.lstProjekteKenn.selectedItems()
        sql_proj = ""
        txt_proj_anzahl = len(projects)
        txt_proj_zaehler = 1
        if txt_proj_anzahl == 0:
            sql_proj = "GZ LIKE '%%'"
        else:
            for project in projects:
                project_split = project.text().split("/")
                project_final = project_split[0].strip()
                if txt_proj_zaehler < txt_proj_anzahl:
                    sql_proj = sql_proj + "GZ LIKE '%" + project_final + "%' OR "
                else:
                    sql_proj = sql_proj + "GZ LIKE '%" + project_final + "%'"
                txt_proj_zaehler = txt_proj_zaehler + 1

        ## Code
        codes = self.ui.lstCode.selectedItems()
        sql_code = ""
        txt_code_anzahl = len(codes)
        txt_code_zaehler = 1
        if txt_code_anzahl == 0:
            sql_code = "Code LIKE '%%'"
        else:
            for code in codes:
                code_split = code.text().split("-")
                code_final = code_split[0].strip()
                if txt_code_zaehler < txt_code_anzahl:
                    sql_code = sql_code + "Code LIKE '" + code_final + "' OR "
                else:
                    sql_code = sql_code + "Code LIKE '" + code_final + "'"
                txt_code_zaehler = txt_code_zaehler + 1

        ## WHERE
        sql_where = "WHERE \n"

        ## SQL Main
        if kennwerteStatusUmbauNeubau == "Neubau":
            sql = sql_main + sql_where + \
                  "(" + sql_codeBez + ") AND \n" + \
                  sql_min + " AND \n" + \
                  sql_max + " AND \n" + \
                  "(" + sql_proj + ") AND \n" + \
                  "(" + sql_code + ") AND \n" + \
                  "[Kennwert Mittelwert] <> 0 AND [Kennwert Mittelwert 3 Bestbieter] <> 0 AND [Kennwert Billigstbieter] <> 0 AND [Kennwert Median] <> 0 AND \n" + \
                  sql_art
        elif kennwerteStatusUmbauNeubau == "Umbau":
            sql = sql_main + sql_where + \
                  "(" + sql_codeBez + ") AND \n" + \
                  sql_min + " AND \n" + \
                  sql_max + " AND \n" + \
                  "(" + sql_proj + ") AND \n" + \
                  "(" + sql_code + ") AND \n" + \
                  "[Kennwert Mittelwert Umbau] <> 0 AND [Kennwert Mittelwert 3 Bestbieter Umbau] <> 0 AND [Kennwert Billigstbieter Umbau] <> 0 AND [Kennwert Median Umbau] <> 0 AND \n" + \
                  sql_art
        else:
            sql = sql_main + sql_where + \
                  "(" + sql_codeBez + ") AND \n" + \
                  sql_min + " AND \n" + \
                  sql_max + " AND \n" + \
                  "(" + sql_proj + ") AND \n" + \
                  "(" + sql_code + ") AND \n" + \
                  "Mittelwert <> 0 AND Mittelwert3Bestbieter <> 0 AND Billigstbieter <> 0 AND Median <> 0 AND \n" + \
                  sql_art

        ## Count Rows
        if kennwerteStatusUmbauNeubau == "Neubau":
            sql_count = """SELECT COUNT(DISTINCT(CONCAT(GZ, Projektname, Objektart, Code, Bezeichnung, Einheit, Bezug, MengeFinal, [Kennwert Mittelwert], [Kennwert Mittelwert 3 Bestbieter], [Kennwert Billigstbieter], [Kennwert Median])))
                FROM KennwerteProjekteFinalView
                """
        elif kennwerteStatusUmbauNeubau == "Umbau":
            sql_count = """SELECT COUNT(DISTINCT(CONCAT(GZ, Projektname, Objektart, Code, Bezeichnung, Einheit, Bezug, MengeFinalUmbau, [Kennwert Mittelwert Umbau], [Kennwert Mittelwert 3 Bestbieter Umbau], [Kennwert Billigstbieter Umbau], [Kennwert Median Umbau])))
                            FROM KennwerteProjekteFinalView
                            """
        else:
            sql_count = """SELECT COUNT(DISTINCT(CONCAT(GZ, Projektname, Objektart, Code, Bezeichnung, Einheit, Bezug, Menge, Mittelwert, Mittelwert3Bestbieter, Billigstbieter, Median)))
                                        FROM KennwerteProjekteFinalView
                                        """

        if kennwerteStatusUmbauNeubau == "Neubau":
            sql_count_final = sql_count + sql_where + \
                              "(" + sql_codeBez + ") AND \n" + \
                              sql_min + " AND \n" + \
                              sql_max + " AND \n" + \
                              "(" + sql_proj + ") AND \n" + \
                              "(" + sql_code + ") AND \n" + \
                              "[Kennwert Mittelwert] <> 0 AND [Kennwert Mittelwert 3 Bestbieter] <> 0 AND [Kennwert Billigstbieter] <> 0 AND [Kennwert Median] <> 0 AND \n" + \
                              sql_art
        elif kennwerteStatusUmbauNeubau == "Umbau":
            sql_count_final = sql_count + sql_where + \
                              "(" + sql_codeBez + ") AND \n" + \
                              sql_min + " AND \n" + \
                              sql_max + " AND \n" + \
                              "(" + sql_proj + ") AND \n" + \
                              "(" + sql_code + ") AND \n" + \
                              "[Kennwert Mittelwert Umbau] <> 0 AND [Kennwert Mittelwert 3 Bestbieter Umbau] <> 0 AND [Kennwert Billigstbieter Umbau] <> 0 AND [Kennwert Median Umbau] <> 0 AND \n" + \
                              sql_art
        else:
            sql_count_final = sql_count + sql_where + \
                              "(" + sql_codeBez + ") AND \n" + \
                              sql_min + " AND \n" + \
                              sql_max + " AND \n" + \
                              "(" + sql_proj + ") AND \n" + \
                              "(" + sql_code + ") AND \n" + \
                              "Mittelwert <> 0 AND Mittelwert3Bestbieter <> 0 AND Billigstbieter <> 0 AND Median <> 0 AND \n" + \
                              sql_art

        return sql, sql_count_final

        # return sql

    def sql_details_kenn(self):
        global kennwerteStatusUmbauNeubau

        gz = self.ui.lblKennDetailGZ.text()
        code = self.ui.lblKennDetailCode.text()

        if kennwerteStatusUmbauNeubau == "Neubau":
            sql_main = """
                        SELECT DISTINCT GZ, Gewerkebezeichnung, Code, Bezug, Bezeichnung, Einheit, FORMAT(Menge, 'N', 'de-de'), FORMAT([Preis Mittelwert], 'N', 'de-de'), FORMAT([Preis Mittelwert 3 Bestbieter], 'N', 'de-de'), FORMAT([Preis Billigstbieter], 'N', 'de-de'), FORMAT([Preis Median], 'N', 'de-de'), FORMAT([Mittelwert], 'N', 'de-de'), FORMAT([Mittelwert 3 Bestbieter], 'N', 'de-de'), FORMAT([Billigstbieter], 'N', 'de-de'), FORMAT([Median], 'N', 'de-de')
                        FROM KennwerteView
                    """
        else:
            sql_main = """
                        SELECT DISTINCT GZ, Gewerkebezeichnung, Code, Bezug, Bezeichnung, Einheit, FORMAT([Menge Umbau], 'N', 'de-de'), FORMAT([Preis Mittelwert Umbau], 'N', 'de-de'), FORMAT([Preis Mittelwert 3 Bestbieter Umbau], 'N', 'de-de'), FORMAT([Preis Billigstbieter Umbau], 'N', 'de-de'), FORMAT([Preis Median Umbau], 'N', 'de-de'), FORMAT([Mittelwert Umbau], 'N', 'de-de'), FORMAT([Mittelwert 3 Bestbieter Umbau], 'N', 'de-de'), FORMAT([Billigstbieter Umbau], 'N', 'de-de'), FORMAT([Median Umbau], 'N', 'de-de')
                        FROM KennwerteView
                    """

        sql_where = "WHERE GZ = '" + gz + "' AND Code LIKE '" + code + "%'"
        if kennwerteStatusUmbauNeubau == "Neubau":
            sql_null = "[Mittelwert] <> 0 AND [Mittelwert 3 Bestbieter] <> 0 AND [Billigstbieter] <> 0 AND [Median] <> 0"
        else:
            sql_null = "[Mittelwert Umbau] <> 0 AND [Mittelwert 3 Bestbieter Umbau] <> 0 AND [Billigstbieter Umbau] <> 0 AND [Median Umbau] <> 0"

        if kennwerteStatusUmbauNeubau == "Neubau":
            orderby = "ORDER BY Code ASC , Bezeichnung ASC, FORMAT(Menge, 'N', 'de-de') DESC;"
        else:
            orderby = "ORDER BY Code ASC , Bezeichnung ASC, FORMAT([Menge Umbau], 'N', 'de-de') DESC;"
        sql = sql_main + sql_where + " AND " + sql_null
        # sql = sql_main + sql_where + orderby
        return sql

    def sql_details_lvpos_kenn(self):
        try:
            rows = self.ui.grdViewDetailKenn.selectionModel().selectedRows()
            if len(rows) == 0:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Kein Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) > 1:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Mehr als eine Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) == 1:
                self.ui.stackedKenn.setCurrentWidget(self.ui.pageDetailLVKenn)
                list_values = []
                for col in range(0, 3):
                    value = self.ui.grdViewDetailKenn.model().index(rows[0].row(), col)
                    data_value = self.ui.grdViewDetailKenn.model().data(value)
                    list_values.append(data_value)

            gz = list_values[0]
            gewerk = list_values[1]
            code = list_values[2]

            sql_pos = """
                                SELECT DISTINCT LVPos, [LVPos Z-Position], [LVPos Stichwort], Code, FORMAT(Menge, 'N', 'de-de'), Einheit, FORMAT([Preis Mittelwert], 'N', 'de-de'), FORMAT([Preis Mittelwert VAL], 'N', 'de-de'), FORMAT([Preis 3 Bestbieter], 'N', 'de-de'), FORMAT([Preis 3 Bestbieter VAL], 'N', 'de-de'), FORMAT([Preis Billigstbieter], 'N', 'de-de'), FORMAT([Preis Billigstbieter VAL], 'N', 'de-de'), FORMAT([Preis Median], 'N', 'de-de'), FORMAT([Preis Median VAL], 'N', 'de-de'), HG, OG
                                FROM GewerkeView
                                WHERE GZ LIKE '%""" + gz + """%' AND
                                        Gewerkebezeichnung LIKE '%""" + gewerk + """%' AND
                                        Code LIKE '""" + code + """%'
                            """
            return sql_pos

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Keine Positionen vorhanden!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def db_fill_table_kenn(self):
        self.setCursor(QCursor(QtCore.Qt.WaitCursor))
        try:
            start_time = datetime.now()

            sql = self.sql_main_kenn()

            model = SqlModelKennwerte()
            model.setQuery(sql[0])
            self.string = str(model.rowCount())

            col = 0

            headers = ['GZ', 'Projektname', 'Objektart', 'Code', 'Bezeichnung', 'Einheit',
                       'Bezug', 'Menge', 'Mittelwert', 'MW 3 Bestbieter',
                       'Billigstbieter', 'Median', 'Status']
            for header in headers:
                model.setHeaderData(col, QtCore.Qt.Horizontal, header)
                col += 1

            view = self.ui.grdViewKenn
            view.setModel(model)
            view.resizeColumnToContents(0)
            view.setColumnWidth(1, 300)
            view.resizeColumnToContents(2)
            view.resizeColumnToContents(3)
            view.setColumnWidth(4, 300)
            view.resizeColumnToContents(5)
            view.resizeColumnToContents(6)
            view.setColumnWidth(7, 75)
            view.setColumnWidth(8, 100)
            view.setColumnWidth(9, 100)
            view.setColumnWidth(10, 100)
            view.setColumnWidth(11, 100)
            view.setColumnWidth(12, 100)

            self.check_screen = CheckScreen()

            end_time = datetime.now()
            diff = (end_time - start_time).total_seconds()
            if model.canFetchMore() == True:
                self.ui.lblDataCount_2.setText(
                    "Abfragedauer: " + self.string + "+ Positionen in " + str(round(diff, 3)) + "s")
            else:
                self.ui.lblDataCount_2.setText(
                    "Abfragedauer: " + self.string + " Positionen in " + str(round(diff, 3)) + "s")

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Fehler bei SQL!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        self.setCursor(QCursor(QtCore.Qt.ArrowCursor))

    def mw_calc_kenn(self):
        list_mittelwert = []
        list_best3 = []
        list_min = []
        list_median = []

        try:
            rows = self.ui.grdViewKenn.selectionModel().selectedRows()
            for row in rows:
                for col in range(8, 12):
                    if col == 8:
                        value = self.ui.grdViewKenn.model().index(row.row(), col)
                        data_value = self.ui.grdViewKenn.model().data(value)
                        list_mittelwert.append(string_to_float(data_value, 2))
                    elif col == 9:
                        value = self.ui.grdViewKenn.model().index(row.row(), col)
                        data_value = self.ui.grdViewKenn.model().data(value)
                        list_best3.append(string_to_float(data_value, 2))
                    elif col == 10:
                        value = self.ui.grdViewKenn.model().index(row.row(), col)
                        data_value = self.ui.grdViewKenn.model().data(value)
                        list_min.append(string_to_float(data_value, 2))
                    elif col == 11:
                        value = self.ui.grdViewKenn.model().index(row.row(), col)
                        data_value = self.ui.grdViewKenn.model().data(value)
                        list_median.append(string_to_float(data_value, 2))

            self.ui.lblMWMittelwertKenn.setText(float_to_string(statistics.mean(list_mittelwert)))
            self.ui.lblMW3BestKenn.setText(float_to_string(statistics.mean(list_best3)))
            self.ui.lblMWBilligstKenn.setText(float_to_string(statistics.mean(list_min)))
            self.ui.lblMWMedianKenn.setText(float_to_string(statistics.mean(list_median)))

        except:
            self.ui.lblMWMittelwertKenn.setText(float_to_string(statistics.mean(0)))
            self.ui.lblMW3BestKenn.setText(float_to_string(statistics.mean(0)))
            self.ui.lblMWBilligstKenn.setText(float_to_string(statistics.mean(0)))
            self.ui.lblMWMedianKenn.setText(float_to_string(statistics.mean(0)))

    def show_kenn_details(self):
        global kennwerteStatusUmbauNeubau
        try:
            self.setCursor(QCursor(QtCore.Qt.WaitCursor))
            rows = self.ui.grdViewKenn.selectionModel().selectedRows()

            if len(rows) == 1:
                self.ui.stackedKenn.setCurrentWidget(self.ui.pageDetailKenn)
                list_values = []
                for col in range(0, 13):
                    value = self.ui.grdViewKenn.model().index(rows[0].row(), col)
                    data_value = self.ui.grdViewKenn.model().data(value)
                    list_values.append(data_value)

                gz = list_values[0]
                projektname = list_values[1]
                code = list_values[3]
                bezeichnung = list_values[4]
                einheit = list_values[5]
                menge = list_values[7]
                bezug = list_values[6]
                kwMittel = list_values[8]
                kwMittel3 = list_values[9]
                kwBilligst = list_values[10]
                kwMedian = list_values[11]
                status = list_values[12]

                self.ui.lblKennDetailBezeichnung.setText(bezeichnung)
                self.ui.lblKennDetailGZ.setText(gz)
                self.ui.lblKennDetailCode.setText(code)
                self.ui.lblKennDetailKWMW.setText(kwMittel)
                self.ui.lblKennDetailBezug.setText(bezug)
                self.ui.lblKennDetailEinheit.setText(einheit)
                self.ui.lblKennDetailEinheit_2.setText(menge)
                self.ui.lblKennDetailKW3Best.setText(kwMittel3)
                self.ui.lblKennDetailKWBilligst.setText(kwBilligst)
                self.ui.lblKennDetailKWMedian.setText(kwMedian)
                self.ui.lblKennDetailProjektname.setText(projektname)
                kennwerteStatusUmbauNeubau = status

                sql = self.sql_details_kenn()
                # print(sql)

                model = SqlModelKennwerteDetails()
                model.setQuery(sql)
                self.string = str(model.rowCount())

                col = 0
                headers = ['GZ', 'Gewerkebezeichnung', 'Code', 'Bezug', 'Bezeichnung', 'Einheit',
                           'Menge', 'Mittelwert', 'MW 3 Bestbieter',
                           'Billigstbieter', 'Median', 'Mittelwert gesamt', 'MW 3 Bestbieter gesamt', 'Billigstbieter gesamt', 'Median gesamt']
                for header in headers:
                    model.setHeaderData(col, QtCore.Qt.Horizontal, header)
                    col += 1

                view = self.ui.grdViewDetailKenn
                view.setModel(model)
                view.resizeColumnToContents(0)
                view.setColumnWidth(1, 300)
                view.resizeColumnToContents(2)
                view.resizeColumnToContents(3)
                view.setColumnWidth(4, 300)
                view.resizeColumnToContents(5)
                view.resizeColumnToContents(6)
                view.setColumnWidth(7, 100)
                view.setColumnWidth(8, 100)
                view.setColumnWidth(9, 100)
                view.setColumnWidth(10, 100)
                view.setColumnWidth(11, 100)
                view.setColumnWidth(12, 100)
                view.setColumnWidth(13, 100)
                view.setColumnWidth(14, 100)

                self.setCursor(QCursor(QtCore.Qt.ArrowCursor))
                self.check_screen = CheckScreen()

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Keine Positionen vorhanden!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def show_kenn_details_lvpos(self):
        try:
            self.setCursor(QCursor(QtCore.Qt.WaitCursor))
            rows = self.ui.grdViewDetailKenn.selectionModel().selectedRows()

            if len(rows) == 0:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Kein Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) > 1:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Mehr als eine Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) == 1:
                self.ui.stackedKenn.setCurrentWidget(self.ui.pageDetailLVKenn)
                list_values = []
                for col in range(0, 11):
                    value = self.ui.grdViewDetailKenn.model().index(rows[0].row(), col)
                    data_value = self.ui.grdViewDetailKenn.model().data(value)
                    list_values.append(data_value)

                gz = self.ui.lblKennDetailGZ.text()
                projektname = self.ui.lblKennDetailProjektname.text()
                code = list_values[2]
                bezeichnung = list_values[4]
                einheit = list_values[5]
                menge = list_values[6]
                bezug = list_values[3]
                kwMittel = list_values[7]
                kwMittel3 = list_values[8]
                kwBilligst = list_values[9]
                kwMedian = list_values[10]

                self.ui.lblKennDetailLVBezeichnung.setText(bezeichnung)
                self.ui.lblKennDetailLVGZ.setText(gz)
                self.ui.lblKennDetailLVCode.setText(code)
                self.ui.lblKennDetailLVKWMW.setText(kwMittel)
                self.ui.lblKennDetailLVBezug.setText(bezug)
                self.ui.lblKennDetailLVEinheit.setText(einheit)
                self.ui.lblKennDetailLVEinheit_2.setText(menge)
                self.ui.lblKennDetailLVKW3Best.setText(kwMittel3)
                self.ui.lblKennDetailLVKWBilligst.setText(kwBilligst)
                self.ui.lblKennDetailLVKWMedian.setText(kwMedian)
                self.ui.lblKennDetailLVProjektname.setText(projektname)

                sql = self.sql_details_lvpos_kenn()

                model = SqlModelKennwerteDetailsLVPos()
                model.setQuery(sql)
                self.string = str(model.rowCount())

                col = 0
                headers = ['LV-Position', 'Z-Pos.', 'Stichwort', 'Code', 'Menge', 'Einheit',
                           'Mittelwert', 'Mittelwert VAL', '3 Bestbieter', '3 Bestbieter VAL',
                           'Billigstbieter', 'Billigstbieter VAL', 'Median', 'Median VAL', 'HG', 'OG']

                for header in headers:
                    model.setHeaderData(col, QtCore.Qt.Horizontal, header)
                    col += 1

                view = self.ui.grdViewDetailLVKenn
                view.setModel(model)
                view.resizeColumnToContents(0)
                view.resizeColumnToContents(1)
                view.resizeColumnToContents(2)
                view.resizeColumnToContents(3)
                view.resizeColumnToContents(4)
                view.resizeColumnToContents(5)
                view.setColumnWidth(6, 100)
                view.setColumnWidth(7, 100)
                view.setColumnWidth(8, 100)
                view.setColumnWidth(9, 100)
                view.setColumnWidth(10, 100)
                view.setColumnWidth(11, 100)
                view.setColumnWidth(12, 100)
                view.setColumnWidth(13, 100)
                view.resizeColumnToContents(14)
                view.resizeColumnToContents(15)

                self.setCursor(QCursor(QtCore.Qt.ArrowCursor))
                self.check_screen = CheckScreen()

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Keine Positionen vorhanden!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    ## Verwaltung
    def set_controls_index(self):
        self.ui.cboIndexQuartal.addItem("1. Quartal")
        self.ui.cboIndexQuartal.addItem("2. Quartal")
        self.ui.cboIndexQuartal.addItem("3. Quartal")
        self.ui.cboIndexQuartal.addItem("4. Quartal")

        sql = """
        SELECT Indexquartal, FORMAT(indexbpi, 'N', 'de-de')
        FROM IndexViewVerwaltung
        ORDER BY year, quarter
        """

        model = SQLModelIndexVerwaltung()
        model.setQuery(sql)

        col = 0
        headers = ['Quartal', 'BPI']
        for header in headers:
            model.setHeaderData(col, QtCore.Qt.Horizontal, header)
            col += 1

        view = self.ui.grdIndex
        view.setModel(model)
        view.setColumnWidth(0, 145)
        view.setColumnWidth(1, 145)

    ##Neue Indizes einfügen
    def index_execute(self):
        index = self.ui.txtIndex.text()
        indexYear = self.ui.txtIndexJahr.text()
        indexYearIsYear = False
        indexQuarterStart = ""
        indexQuarterEnd = ""
        indexQuarterList = []

        index = string_to_float(index, 1)
        indexYear = indexYear.strip()

        if (len(indexYear) == 4):
            indexYearIsYear = True

        if (indexYearIsYear == True):
            if (self.ui.cboIndexQuartal.currentIndex() == 0 and indexYearIsYear == True):
                indexQuarterStart = indexYear + "0101"
                indexQuarterEnd = indexYear + "0331"
            elif (self.ui.cboIndexQuartal.currentIndex() == 1 and indexYearIsYear == True):
                indexQuarterStart = indexYear + "0401"
                indexQuarterEnd = indexYear + "0630"
            elif (self.ui.cboIndexQuartal.currentIndex() == 2 and indexYearIsYear == True):
                indexQuarterStart = indexYear + "0701"
                indexQuarterEnd = indexYear + "0930"
            elif (self.ui.cboIndexQuartal.currentIndex() == 3 and indexYearIsYear == True):
                indexQuarterStart = indexYear + "1001"
                indexQuarterEnd = indexYear + "1231"

            for dt in rrule.rrule(rrule.DAILY, dtstart=datetime.strptime(indexQuarterStart, '%Y%m%d'), until=datetime.strptime(indexQuarterEnd, '%Y%m%d')):
                tempList = [dt.strftime('%Y-%m-%d %H:%M:%S'), index]
                indexQuarterList.append(tempList)

            self.query = QtSql.QSqlQuery()

            ## Insert all available Indizes
            for item in indexQuarterList:
                datumTemp = str(item[0])
                indexTemp = float(item[1])
                self.qry_insert_bpi = "INSERT INTO [PreisDB].[db_owner].[tblIndezes] (indexinletIDRef, indexBPI, indexDatum) VALUES ('1', '"
                self.qry_insert_bpi += str(indexTemp)
                self.qry_insert_bpi += """', '""" + datumTemp + "');"
                self.query.exec(self.qry_insert_bpi)

            ## Update all available Indizes
            for item in indexQuarterList:
                datumTemp = str(item[0])
                indexTemp = float(item[1])
                self.qry_insert_bpi = """UPDATE [PreisDB].[db_owner].[tblIndezes] SET indexBPI = '""" + str(indexTemp) + """' WHERE indexDatum = '""" + str(
                    datumTemp) + """';"""
                self.query.exec(self.qry_insert_bpi)

            msgIndex = QMessageBox()
            msgIndex.setWindowIcon(QtGui.QIcon(company_icon))
            msgIndex.setIcon(QMessageBox.Information)
            msgIndex.setWindowFlag(Qt.WindowStaysOnTopHint)
            msgIndex.setText("Die neuen Indizes wurden eingefügt!")
            msgIndex.setWindowTitle("Preisdatenbank")
            msgIndex.setStandardButtons(QMessageBox.Ok)
            msgIndex.exec_()

        else:
            msgIndex = QMessageBox()
            msgIndex.setWindowIcon(QtGui.QIcon(company_icon))
            msgIndex.setIcon(QMessageBox.Warning)
            msgIndex.setWindowFlag(Qt.WindowStaysOnTopHint)
            msgIndex.setText("Das Jahr ist ungültig!")
            msgIndex.setWindowTitle("Preisdatenbank")
            msgIndex.setStandardButtons(QMessageBox.Ok)
            msgIndex.exec_()


class IndexRechner(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(IndexRechner, self).__init__(parent, QtCore.Qt.WindowStaysOnTopHint)

        self.ui = Ui_IndexRechner()
        self.ui.setupUi(self)
        self.partnerDialogCalc = parent

        self.setWindowTitle("Indexrechner")
        # self.setWindowIcon(QtGui.QIcon(".icon\\favicon.ico"))
        self.setWindowIcon(QtGui.QIcon(company_icon))

        self.load_index()

        self.ui.cmdRefresh.clicked.connect(self.reset_index)
        self.ui.cmdCalc.clicked.connect(self.index_values)
        self.ui.cmdImport.clicked.connect(self.import_preis)

    def reset_index(self):
        self.load_index()
        self.ui.lblIndexVon.setText("-")
        self.ui.lblIndexBis.setText("-")
        self.ui.lblPreissteigerung.setText("-")

    def load_index(self):
        ## Anlegen des Query-Objekts
        self.query = QtSql.QSqlQuery()

        ##Abfrage Preisbasis
        self.qry_pb = """
            SELECT MIN(Datum), MAX(Datum)
            FROM IndexView
        """
        self.query.exec(self.qry_pb)

        self.query.first()
        low = self.query.value(0).toPyDateTime()
        high = self.query.value(1).toPyDateTime()

        self.ui.datIndexVon.setDateRange(QDate(low.year, low.month, low.day), QDate(high.year, high.month, high.day))
        self.ui.datIndexBis.setDateRange(QDate(low.year, low.month, low.day), QDate(high.year, high.month, high.day))
        self.ui.datIndexVon.setDate(QDate(low.year, low.month, low.day))
        self.ui.datIndexBis.setDate(QDate(high.year, high.month, high.day))

    def import_preis(self):
        try:
            rows = self.partnerDialogCalc.ui.grdView.selectionModel().selectedRows()
            if len(rows) == 0:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Kein Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) > 1:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon(company_icon))
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowFlag(Qt.WindowStaysOnTopHint)
                msg.setText("Mehr als eine Position gewählt!")
                msg.setWindowTitle("Preisdatenbank")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
                return

            if len(rows) == 1:
                list_values = []
                for col in range(8, 19):
                    if col == 8:
                        value = self.partnerDialogCalc.ui.grdView.model().index(rows[0].row(), col)
                        data_value = self.partnerDialogCalc.ui.grdView.model().data(value)
                        date_fin = datetime.strptime(data_value, '%Y-%m-%d')
                        list_values.append(date_fin)

                    if col >= 15 and col <= 18:
                        value = self.partnerDialogCalc.ui.grdView.model().index(rows[0].row(), col)
                        data_value = self.partnerDialogCalc.ui.grdView.model().data(value)
                        value_fin = string_to_float(data_value, 2)
                        list_values.append(value_fin)

                self.ui.lblMittelwert.setText(format_decimal(list_values[1], format='#,##0.00', locale='de_DE'))
                self.ui.lbl3Best.setText(format_decimal(list_values[2], format='#,##0.00', locale='de_DE'))
                self.ui.lblBilligst.setText(format_decimal(list_values[3], format='#,##0.00', locale='de_DE'))
                self.ui.lblMedian.setText(format_decimal(list_values[4], format='#,##0.00', locale='de_DE'))
                self.ui.datIndexVon.setDate(QDate(list_values[0].year, list_values[0].month, list_values[0].day))

                self.index_values()

                von = string_to_float(self.ui.lblIndexVon.text(), 2)
                bis = string_to_float(self.ui.lblIndexBis.text(), 2)
                preissteigerung = float(bis / von)

                self.ui.lblMittelwertVAL.setText(format_decimal(list_values[1] * preissteigerung, format='#,##0.00', locale='de_DE'))
                self.ui.lbl3BestVAL.setText(format_decimal(list_values[2] * preissteigerung, format='#,##0.00', locale='de_DE'))
                self.ui.lblBilligstVAL.setText(format_decimal(list_values[3] * preissteigerung, format='#,##0.00', locale='de_DE'))
                self.ui.lblMedianVAL.setText(format_decimal(list_values[4] * preissteigerung, format='#,##0.00', locale='de_DE'))

        except:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon(company_icon))
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowFlag(Qt.WindowStaysOnTopHint)
            msg.setText("Keine Positionen vorhanden!")
            msg.setWindowTitle("Preisdatenbank")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

        # table = self.partnerDialog.ui.grdAbfrage
        # row = table.currentRow()
        # if not row == -1:
        #     mittelwert = table.item(row, 3).text()
        #     mittelwert3best = table.item(row, 5).text()
        #     billigst = table.item(row, 7).text()
        #     median = table.item(row, 9).text()
        #     date = table.item(row, 11).text()
        #
        #     print(date)
        #     print(type(date))
        #     date_fin = datetime.strptime(date, '%d-%m-%Y')
        #     mittelwert_calc = round(float(mittelwert.replace(".", "").replace(",", ".")), 2)
        #     mittelwert3best_calc = round(float(mittelwert3best.replace(".", "").replace(",", ".")), 2)
        #     billigst_calc = round(float(billigst.replace(".", "").replace(",", ".")), 2)
        #     median_calc = round(float(median.replace(".", "").replace(",", ".")), 2)
        #     self.ui.lblMittelwert.setText(format_decimal(mittelwert_calc, format='#,##0.00', locale='de_DE'))
        #     self.ui.lbl3Best.setText(format_decimal(mittelwert3best_calc, format='#,##0.00', locale='de_DE'))
        #     self.ui.lblBilligst.setText(format_decimal(billigst_calc, format='#,##0.00', locale='de_DE'))
        #     self.ui.lblMedian.setText(format_decimal(median_calc, format='#,##0.00', locale='de_DE'))
        #
        #     self.ui.datIndexVon.setDate(QDate(date_fin.year, date_fin.month, date_fin.day))
        #
        #     self.index_values()
        #
        #     von = float(self.ui.lblIndexVon.text())
        #     bis = float(self.ui.lblIndexBis.text())
        #     preissteigerung = float(bis/von)
        #
        #     self.ui.lblMittelwertVAL.setText(format_decimal(mittelwert_calc*preissteigerung, format='#,##0.00', locale='de_DE'))
        #     self.ui.lbl3BestVAL.setText(format_decimal(mittelwert3best_calc*preissteigerung, format='#,##0.00', locale='de_DE'))
        #     self.ui.lblBilligstVAL.setText(format_decimal(billigst_calc*preissteigerung, format='#,##0.00', locale='de_DE'))
        #     self.ui.lblMedianVAL.setText(format_decimal(median_calc*preissteigerung, format='#,##0.00', locale='de_DE'))

    def index_values(self):
        ## Anlegen des Query-Objekts
        self.query = QtSql.QSqlQuery()

        ##Abfrage Preisbasis
        self.qry_pb = """
            SELECT  MIN(Preisbasis), MAX(Preisbasis)
            FROM GewerkeView
        """
        self.query.exec(self.qry_pb)

        self.query.first()
        low = self.query.value(0).toPyDateTime()
        high = self.query.value(1).toPyDateTime()

        txt_start = self.ui.datIndexVon.text()
        start = datetime.strptime(txt_start, '%d.%m.%Y')

        txt_ende = self.ui.datIndexBis.text()
        ende = datetime.strptime(txt_ende, '%d.%m.%Y')

        self.qry_min = """
            SELECT FORMAT(BPI, 'N', 'de-de')
            FROM IndexView
            WHERE Datum = '""" + str(start.date()) + """'
        """
        self.query.exec(self.qry_min)
        self.query.first()
        self.ui.lblIndexVon.setText(self.query.value(0))

        self.qry_max = """
            SELECT FORMAT(BPI, 'N', 'de-de')
            FROM IndexView
            WHERE Datum = '""" + str(ende.date()) + """'
        """
        self.query.exec(self.qry_max)
        self.query.first()
        self.ui.lblIndexBis.setText(self.query.value(0))

        von = string_to_float(self.ui.lblIndexVon.text(), 2)
        bis = string_to_float(self.ui.lblIndexBis.text(), 2)

        preissteigerung = round(float(bis / von), 3)
        self.ui.lblPreissteigerung.setText(str(preissteigerung))

        mittelwert_calc = round(float(self.ui.lblMittelwert.text().replace(".", "").replace(",", ".")), 2)
        mittelwert3best_calc = round(float(self.ui.lbl3Best.text().replace(".", "").replace(",", ".")), 2)
        billigst_calc = round(float(self.ui.lblBilligst.text().replace(".", "").replace(",", ".")), 2)
        median_calc = round(float(self.ui.lblMedian.text().replace(".", "").replace(",", ".")), 2)
        self.ui.lblMittelwertVAL.setText(format_decimal(round(mittelwert_calc * preissteigerung, 2), format='#,##0.00', locale='de_DE'))
        self.ui.lbl3BestVAL.setText(format_decimal(round(mittelwert3best_calc * preissteigerung, 2), format='#,##0.00', locale='de_DE'))
        self.ui.lblBilligstVAL.setText(format_decimal(round(billigst_calc * preissteigerung, 2), format='#,##0.00', locale='de_DE'))
        self.ui.lblMedianVAL.setText(format_decimal(round(median_calc * preissteigerung, 2), format='#,##0.00', locale='de_DE'))


## ==> APP FUNCTIONS

splashcreen = SplashScreen()
splashcreen.show()

sys.exit(app.exec_())
